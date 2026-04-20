from __future__ import annotations

import openpyxl
import requests
import re
import csv
import io

from pathlib import Path
from typing import Optional, List, Tuple, Set, Dict
from urllib.parse import urlparse
from bs4 import BeautifulSoup

from dataclass import SiteConfig
from helper import clean_text, get_parser
from web.func import create_session
from config import Defaults, Formats, Settings


def infer_extensions_from_formats(formats_text: str) -> Tuple[str, ...]:
    found: Set[str] = set()
    for raw_token in formats_text.split(","):
        token = clean_text(raw_token).lower()
        if token in Formats.ignored_format_tokens:
            continue
        if token.startswith(".") and len(token) > 1:
            found.add(token)
            continue
        if token in Formats.extension_map:
            found.update(Formats.extension_map[token])
            continue
        normalized = token.replace(" ", "").replace("-", "")
        for key, exts in Formats.extension_map.items():
            key_norm = key.replace(" ", "").replace("-", "")
            if normalized == key_norm:
                found.update(exts)
                break
    return tuple(sorted(found))


def build_site_config(
        category: str,
        repository: str,
        start_url: str,
        common_formats: str) -> Optional[SiteConfig]:
    
    category = clean_text(category)
    repository = clean_text(repository)
    start_url = clean_text(start_url)
    common_formats = clean_text(common_formats)

    if not repository or not start_url:
        return None

    extensions = infer_extensions_from_formats(common_formats)
    if not extensions:
        return None

    return SiteConfig(
        category=category,
        repository=repository,
        start_url=start_url,
        extensions=extensions,
    )


def load_site_configs(
    workbook: Optional[str],
    sheet_name: str,
    config_url: Optional[str],
    gid: Optional[str],
    session: requests.Session) -> List[SiteConfig]:

    if config_url:
        if not is_google_sheet_url(config_url):
            raise ValueError(
                "Currently only public Google Sheet URLs are supported for --config-url.")
        return load_site_configs_from_google_sheet(
            config_url,
            sheet_name=sheet_name,
            gid=gid,
            session=session,
        )

    if not workbook:
        raise ValueError("Either --workbook or --config-url must be provided.")

    if re.match(r"^https?://", workbook, flags=re.I):
        if not is_google_sheet_url(workbook):
            raise ValueError(
                "Remote config URLs are currently supported only for public Google Sheets.")
        return load_site_configs_from_google_sheet(
            workbook,
            sheet_name=sheet_name,
            gid=gid,
            session=session,
        )

    return load_site_configs_from_excel(Path(workbook), sheet_name)


def parse_config_rows(rows: List[List[object]]) -> List[SiteConfig]:
    configs: List[SiteConfig] = []
    for row in rows[1:]:  # skip header
        category = row[0] if len(row) > 0 else ""
        repository = row[1] if len(row) > 1 else ""
        start_url = row[2] if len(row) > 2 else ""
        common_formats = row[5] if len(row) > 5 else ""
        config = build_site_config(category, repository, start_url, common_formats)
        if config is not None:
            configs.append(config)
    return configs


def load_site_configs_from_excel(workbook_path: Path,
                                 sheet_name: str) -> List[SiteConfig]:
    
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook. "
                         "Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return parse_config_rows(rows)


def is_google_sheet_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.netloc.lower() in {"docs.google.com", "drive.google.com"} \
        and "/spreadsheets/" in parsed.path.lower()


def extract_google_sheet_id(url: str) -> Optional[str]:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return match.group(1) if match else None


def load_site_configs_from_google_sheet(
    sheet_url: str,
    sheet_name: str,
    gid: str,
    session: Optional[requests.Session] = None) -> List[SiteConfig]:

    local_session = session or create_session()
    sheet_id = extract_google_sheet_id(sheet_url)

    if not sheet_id:
        raise ValueError("Could not extract Google Sheet ID from URL.")

    # Priority: explicit gid > sheet name > default
    if gid:
        resolved_gid = gid
    elif sheet_name:
        try:
            resolved_gid = resolve_google_sheet_gid_by_name(
                sheet_url=sheet_url,
                target_sheet_name=sheet_name,
                session=local_session,
            )
        except Exception as e:
            print(f"[warn] Could not resolve sheet '{sheet_name}' → "
                  "falling back to default tab (gid=0). "
                  f"Reason: {e}")
            resolved_gid = "0"
    else:
        resolved_gid = "0"
    
    export_url = Settings.export_url.format(sheet_id = sheet_id,
                                            resolved_gid = resolved_gid)
    response = local_session.get(export_url,
                                 timeout=Defaults.request_timeout,
                                 allow_redirects=True)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()
    if "text/csv" not in content_type and "application/vnd.ms-excel" not in content_type:
        preview = response.text[:200].strip().replace("\n", " ")
        raise ValueError(
            "Google Sheet did not return CSV. "
            "It may not be public, or the requested tab may not be accessible. "
            f"Content-Type={content_type!r}, preview={preview!r}"
        )

    decoded = response.content.decode(response.encoding or "utf-8", errors="replace")
    reader = csv.reader(io.StringIO(decoded))
    rows = [list(row) for row in reader]
    return parse_config_rows(rows)


def resolve_google_sheet_gid_by_name(
    sheet_url: str,
    target_sheet_name: str,
    session: requests.Session) -> str:

    """
    For a public Google Sheet, fetch the spreadsheet page and map visible tab names to gid values.
    Ignores any gid present in the input URL.
    """
    response = session.get(sheet_url,
                           timeout=Defaults.request_timeout,
                           allow_redirects=True)
    
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()
    if "html" not in content_type:
        raise ValueError("Google Sheet URL did not return HTML, "
                         "so tab metadata could not be inspected.")

    html = response.text
    soup = BeautifulSoup(html, get_parser())

    wanted = clean_text(target_sheet_name).casefold()
    if not wanted:
        raise ValueError("A non-empty --sheet value is required for Google Sheets input.")

    # Google often renders sheet tab links as anchors containing gid in the href.
    candidates: List[Tuple[str, str]] = []

    for a in soup.find_all("a", href=True):
        href = str(a["href"])
        text = clean_text(a.get_text(" ", strip=True))
        if "gid=" not in href:
            continue

        gid_match = re.search(r"(?:[?#&])gid=([0-9]+)", href)
        if gid_match and text:
            candidates.append((text, gid_match.group(1)))

    # Exact case-insensitive title match
    for title, gid in candidates:
        if clean_text(title).casefold() == wanted:
            return gid

    # Helpful fallback error
    visible_tabs = [title for title, _ in candidates]
    raise ValueError(
        f"Could not find a Google Sheet tab named {target_sheet_name!r}. "
        f"Visible parsed tabs: {visible_tabs}")


def write_csv(rows: List[Dict[str, str]],
              output_csv: Path) -> None:

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "Dataset Name",
        "Category",
        "Size",
        "Data Storage Type",
        "Accessible?",
        "Download Path",
        "Document Path",
        "Extracted Paths",
        "Contained File Types",
    ]
    with open(output_csv, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
